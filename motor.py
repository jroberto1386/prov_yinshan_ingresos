# ─────────────────────────────────────────────
#  motor.py  —  Parseo de XMLs CFDI 4.0 y
#               generación de layout CONTPAq
# ─────────────────────────────────────────────

import zipfile
import io
from datetime import datetime
from xml.etree import ElementTree as ET

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from config import (
    RFC_EMISOR, CTA_INGRESOS, CTA_IVA_TRAS,
    TIPO_POL, ID_DIARIO, CATALOGO_CLIENTES
)

# ── Namespaces CFDI 4.0 ──────────────────────
NS = {
    "cfdi": "http://www.sat.gob.mx/cfd/4",
    "tfd":  "http://www.sat.gob.mx/TimbreFiscalDigital",
}

# ── Encabezados fijos del layout CONTPAq (22 filas) ──
HEADERS = [
    # Fila 1 – Egreso(EG)
    ["Egreso(EG)", "IdDocumentoDe", "TipoDocumento", "Folio", "Fecha", "FechaAplicacion",
     "CodigoPersona", "BeneficiarioPagador", "IdCuentaCheques", "CodigoMoneda",
     "Total", "Referencia", "Origen", "BancoDestino", "CuentaDestino",
     "OtroMetodoDePago", "Guid", None, None, "TipoCambio",
     "UUIDRep", "NodoPago", "CodigoMonedaTipoCambio", "NumAsoc"],
    # Fila 2 – deposito.1(DE)
    ["deposito.1(DE)", "IdDocumentoDe", "TipoDocumento", "Folio", "Fecha", "Ejercicio",
     "Periodo", "FechaAplicacion", "EjercicioAp", "PeriodoAp",
     "IdCuentaCheques", "NatBancaria", "Naturaleza", "Total", "Referencia",
     "Concepto", "EsConciliado", "IdMovEdoCta", "EjercicioPol", "PeriodoPol",
     "TipoPol", "NumPol", "FormaDeposito", "IdPoliza", "Origen",
     "IdDocumento", "PolizaAgrupada", "UsuarioCrea", "UsuarioModifica", "tieneCFD", "Guid"],
    # Fila 3 – ingreso.1(IN)
    ["ingreso.1(IN)", "IdDocumentoDe", "TipoDocumento", "Folio", "Fecha", "FechaAplicacion",
     "CodigoPersona", "BeneficiarioPagador", "IdCuentaCheques", "CodigoMoneda",
     "Total", "Referencia", "Origen", "BancoOrigen", "CuentaOrigen",
     "OtroMetodoDePago", "Guid", None, None, "TipoCambio",
     "NumeroCheque", "UUIDRep", "NodoPago", "CodigoMonedaTipoCambio", "NumAsoc"],
    # Fila 4 – Datos para CONTPAQi Factura Electrónica®(FE)
    ["Datos para CONTPAQi Factura Electrónica®(FE)", "RutaAnexo", "ArchivoAnexo"],
    # Fila 5 – Movimiento de póliza(M1)
    ["Movimiento de póliza(M1)", "IdCuenta", "Referencia", "TipoMovto", "Importe",
     "IdDiario", "ImporteME", "Concepto", "IdSegNeg", "Guid", "FechaAplicacion"],
    # Fila 6 – Devolución de IVA (IETU)(W)
    ["Devolución de IVA (IETU)(W)", "IETUDeducible", "IETUModificado"],
    # Fila 7 – Devolución de IVA(V)
    ["Devolución de IVA(V)", "IdProveedor", "ImpTotal", "PorIVA", "ImpBase",
     "ImpIVA", "CausaIVA", "ExentoIVA", "Serie", "Folio", "Referencia",
     "OtrosImptos", "ImpSinRet", "IVARetenido", "ISRRetenido", "GranTotal",
     "EjercicioAsignado", "PeriodoAsignado", "IdCuenta", "IVAPagNoAcred",
     "UUID", None, "IEPS"],
    # Fila 8 – Asociación de nodo de pago(AP)
    ["Asociación de nodo de pago(AP)", "UUIDRep", "NumNodoPago", "GuidReferencia", "AplicationType"],
    # Fila 9 – Periodo de causación de IVA(R)
    ["Periodo de causación de IVA(R)", "EjercicioAsignado", "PeriodoAsignado"],
    # Fila 10 – Póliza(P)
    ["Póliza(P)", "Fecha", "TipoPol", "Folio", "Clase", "IdDiario",
     "Concepto", "SistOrig", "Impresa", "Ajuste", "Guid"],
    # Fila 11 – Asociación movimiento(AM)
    ["Asociación movimiento(AM)", "UUID"],
    # Fila 12 – Comprobantes(MC)
    ["Comprobantes(MC)", "IdCuentaFlujoEfectivo", "IdSegmentoNegCtaFlujo", "Fecha",
     "Serie", "Folio", "UUID", "ClaveRastreo", "Referencia", "IdProveedor",
     "CodigoConceptoIETU", "ImpNeto", "ImpNetoME", "IdCuentaNeto",
     "IdSegmentoNegNeto", "PorIVA", "ImporteIVA", "ImporteIVAME",
     "IVATasaExcenta", "IdCuentaIVA", "IdSegmentoNegIVA", "NombreImpuesto",
     "ImpImpuesto", "ImpImpuestoME", "IdCuentaImpuesto", "IdSegmentoNegImp",
     "ImpOtrosGastos", "ImpOtrosGastosME", "IdCuentaOtrosGastos",
     "IdSegmentoNegOtrosGastos", "IVARetenido", "IVARetenidoME",
     "IdCuentaRetIVA", "IdSegmentoNegRetIVA", "ISRRetenido", "ISRRetenidoME",
     "IdCuentaRetISR", "IdSegmentoNegRetISR", "NombreOtrasRetenciones",
     "ImpOtrasRetenciones", "ImpOtrasRetencionesME", "IdCuentaOtrasRetenciones",
     "IdSegmentoNegOtrasRet", "BaseIVADIOT", "BaseIETU", "IVANoAcreditable",
     "ImpTotalErogacion", "IVAAcreditable", "ImpExtra1", "ImpExtra2",
     "IdCategoria", "IdSubCategoria", "TipoCambio", "IdDocGastos",
     "EsCapturaCompleta", "FolioStr"],
    # Fila 13 – Movimiento de póliza(M)
    ["Movimiento de póliza(M)", "IdCuenta", "Referencia", "TipoMovto", "Importe",
     "IdDiario", "ImporteME", "Concepto", "IdSegNeg"],
    # Fila 14 – Dispersiones de pago(DP)
    ["Dispersiones de pago(DP)", "UUID", "UUIDRep", "GuidRef", "NumNodoPago",
     "FechaPago", "TotalPago", "TipoCambio", "TotalPagoComprobante"],
    # Fila 15 – Devolución de IVA (IETU)(W2)
    ["Devolución de IVA (IETU)(W2)", "IETUDeducible", "IETUAcreditable",
     "IETUModificado", "IdConceptoIETU"],
    # Fila 16 – Movimientos de impuestos(I)
    ["Movimientos de impuestos(I)", "IdPersona", "EjercicioAsignado",
     "PeriodoAsignado", "IdCuenta", "AplicaImpuesto", "Serie", "Folio",
     "Referencia", "UUID", "Origen", "Computable", "TipoMovimiento",
     "TipoFactor", "Impuesto", "ObjetoImpuesto", "NombreImpLocal",
     "TasaOCuota", "ImpBase", "ImpImpuesto", "ImpTotal", "Desglosado",
     "IVANoAcred", "AcumulaIETU", "IdConceptoIETU", "IETUDeducible",
     "IETUModificado", "IETUAcreditable", "GuidMov", "GuidMovPadre",
     "Migrado", "ConceptoIVA", "SubconceptoIVA", "ClasificadorIVA",
     "ProporcionDIOT", "DeducibleDIOT"],
    # Fila 17 – Asociación documento(AD)
    ["Asociación documento(AD)", "UUID"],
    # Fila 18 – Cheque(CH)
    ["Cheque(CH)", "IdDocumentoDe", "TipoDocumento", "Folio", "Fecha",
     "FechaAplicacion", "CodigoPersona", "BeneficiarioPagador",
     "IdCuentaCheques", "CodigoMoneda", "Total", "Referencia", "Origen",
     "CuentaDestino", "BancoDestino", "Guid", None, "OtroMetodoDePago",
     "TipoCambio", "UUIDRep", "NodoPago", "CodigoMonedaTipoCambio", "NumAsoc"],
    # Fila 19 – IngresosNoDepositados.1(DI)
    ["IngresosNoDepositados.1(DI)", "IdDocumentoDe", "TipoDocumento", "Folio",
     "Fecha", "Ejercicio", "Periodo", "FechaAplicacion", "EjercicioAp",
     "PeriodoAp", "CodigoPersona", "BeneficiarioPagador", "NatBancaria",
     "Naturaleza", "CodigoMoneda", "CodigoMonedaTipoCambio", "TipoCambio",
     "Total", "Referencia", "Concepto", "EsAsociado", "UsuAutorizaPresupuesto",
     "PosibilidadPago", "EsProyectado", "Origen", "IdChequeOrigen",
     "TipoCambioDeposito", "IdDocumento", "EsAnticipo", "EsTraspasado",
     "UsuarioCrea", "UsuarioModifica", "tieneCFD", "Guid",
     "CuentaOrigen", "BancoOrigen", "OtroMetodoDePago", "NumeroCheque", "NumAsoc"],
    # Fila 20 – Causación de IVA (Concepto de IETU)(E)
    ["Causación de IVA (Concepto de IETU)(E)", "IdConceptoIETU"],
    # Fila 21 – Causación de IVA (IETU)(D)
    ["Causación de IVA (IETU)(D)", "IVATasa15NoAcred", "IVATasa10NoAcred",
     "IETU", "Modificado", "Origen", "TotTasa16", "BaseTasa16", "IVATasa16",
     "IVATasa16NoAcred", "TotTasa11", "BaseTasa11", "IVATasa11",
     "IVATasa11NoAcred", "TotTasa8", "BaseTasa8", "IVATasa8", "IVATasa8NoAcred"],
    # Fila 22 – Causación de IVA(C)
    ["Causación de IVA(C)", "Tipo", "TotTasa15", "BaseTasa15", "IVATasa15",
     "TotTasa10", "BaseTasa10", "IVATasa10", "TotTasa0", "BaseTasa0",
     "TotTasaExento", "BaseTasaExento", "TotOtraTasa", "BaseOtraTasa",
     "IVAOtraTasa", "ISRRetenido", "TotOtros", "IVARetenido",
     "Captado", "NoCausar", "IEPS"],
]


# ── Parseo de un XML ──────────────────────────
def parsear_xml(xml_bytes):
    """
    Lee un XML CFDI 4.0 y devuelve un dict con los campos relevantes,
    o None si no es de tipo I (ingreso) o si el emisor no coincide.
    """
    try:
        root = ET.fromstring(xml_bytes)
    except ET.ParseError:
        return None

    # Solo facturas de ingreso
    tipo = root.get("TipoDeComprobante", "")
    if tipo != "I":
        return None

    # Verificar RFC emisor
    emisor = root.find("cfdi:Emisor", NS)
    if emisor is None or emisor.get("Rfc", "") != RFC_EMISOR:
        return None

    # Datos del receptor
    receptor = root.find("cfdi:Receptor", NS)
    if receptor is None:
        return None
    rfc_receptor = receptor.get("Rfc", "")

    # Validar que el RFC esté en catálogo
    if rfc_receptor not in CATALOGO_CLIENTES:
        return None  # RFC desconocido → omitir (registrar en warnings)

    cliente = CATALOGO_CLIENTES[rfc_receptor]

    # Campos del comprobante
    fecha_str = root.get("Fecha", "")       # formato: YYYY-MM-DDTHH:MM:SS
    folio     = root.get("Folio", "")
    serie     = root.get("Serie", "")
    subtotal  = float(root.get("SubTotal", "0") or 0)
    total     = float(root.get("Total", "0") or 0)

    # IVA trasladado (suma de todos los traslados)
    iva = 0.0
    impuestos = root.find("cfdi:Impuestos", NS)
    if impuestos is not None:
        traslados = impuestos.find("cfdi:Traslados", NS)
        if traslados is not None:
            for traslado in traslados.findall("cfdi:Traslado", NS):
                iva += float(traslado.get("Importe", "0") or 0)

    # UUID del TimbreFiscalDigital
    complemento = root.find("cfdi:Complemento", NS)
    uuid = ""
    if complemento is not None:
        tfd = complemento.find("tfd:TimbreFiscalDigital", NS)
        if tfd is not None:
            uuid = tfd.get("UUID", "").upper()

    # Parsear fecha
    try:
        fecha = datetime.strptime(fecha_str[:10], "%Y-%m-%d")
    except ValueError:
        fecha = None

    # Folio de referencia: Serie-Folio o solo Folio
    ref_folio = f"{serie}{folio}".strip() if serie else folio

    return {
        "fecha":       fecha,
        "folio":       ref_folio,
        "rfc_receptor": rfc_receptor,
        "nombre_cliente": cliente["nombre"],
        "cta_cxc":     cliente["cuenta"],
        "subtotal":    round(subtotal, 2),
        "iva":         round(iva, 2),
        "total":       round(total, 2),
        "uuid":        uuid,
    }


# ── Procesamiento del ZIP ─────────────────────
def procesar_zip(zip_bytes, num_poliza_inicio, callback_progreso=None):
    """
    Recibe los bytes del ZIP, procesa todos los XMLs y devuelve:
      - rows: lista de dicts con datos por factura
      - omitidos: lista de nombres de archivos omitidos
      - sin_catalogo: list of (filename, rfc) sin mapeo
    """
    facturas     = []
    omitidos     = []
    sin_catalogo = []

    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        nombres = [n for n in zf.namelist() if n.lower().endswith(".xml")]
        total   = len(nombres)

        for i, nombre in enumerate(nombres):
            if callback_progreso:
                callback_progreso(i + 1, total)

            with zf.open(nombre) as f:
                xml_bytes = f.read()

            resultado = parsear_xml_con_warning(xml_bytes)

            if resultado is None:
                omitidos.append(nombre)
            elif resultado == "sin_catalogo":
                sin_catalogo.append(nombre)
            else:
                facturas.append(resultado)

    # Ordenar por fecha
    facturas.sort(key=lambda x: (x["fecha"] or datetime.min, x["folio"]))

    return facturas, omitidos, sin_catalogo


def parsear_xml_con_warning(xml_bytes):
    """
    Igual que parsear_xml pero distingue entre
    'omitido por tipo/emisor' y 'sin RFC en catálogo'.
    """
    try:
        root = ET.fromstring(xml_bytes)
    except ET.ParseError:
        return None

    tipo = root.get("TipoDeComprobante", "")
    if tipo != "I":
        return None  # P, E, N, etc.

    emisor = root.find("cfdi:Emisor", NS)
    if emisor is None or emisor.get("Rfc", "") != RFC_EMISOR:
        return None

    receptor = root.find("cfdi:Receptor", NS)
    if receptor is None:
        return None
    rfc_receptor = receptor.get("Rfc", "")

    if rfc_receptor not in CATALOGO_CLIENTES:
        return "sin_catalogo"

    cliente = CATALOGO_CLIENTES[rfc_receptor]

    fecha_str = root.get("Fecha", "")
    folio     = root.get("Folio", "")
    serie     = root.get("Serie", "")
    subtotal  = float(root.get("SubTotal", "0") or 0)
    total     = float(root.get("Total", "0") or 0)

    iva = 0.0
    impuestos = root.find("cfdi:Impuestos", NS)
    if impuestos is not None:
        traslados = impuestos.find("cfdi:Traslados", NS)
        if traslados is not None:
            for traslado in traslados.findall("cfdi:Traslado", NS):
                iva += float(traslado.get("Importe", "0") or 0)

    complemento = root.find("cfdi:Complemento", NS)
    uuid = ""
    if complemento is not None:
        tfd = complemento.find("tfd:TimbreFiscalDigital", NS)
        if tfd is not None:
            uuid = tfd.get("UUID", "").upper()

    try:
        fecha = datetime.strptime(fecha_str[:10], "%Y-%m-%d")
    except ValueError:
        fecha = None

    ref_folio = f"{serie}{folio}".strip() if serie else folio

    return {
        "fecha":          fecha,
        "folio":          ref_folio,
        "rfc_receptor":   rfc_receptor,
        "nombre_cliente": cliente["nombre"],
        "cta_cxc":        cliente["cuenta"],
        "subtotal":       round(subtotal, 2),
        "iva":            round(iva, 2),
        "total":          round(total, 2),
        "uuid":           uuid,
    }


# ── Generación del Excel CONTPAq ──────────────
def generar_excel(facturas, num_poliza_inicio):
    """
    Recibe lista de facturas y el número de póliza inicial.
    Devuelve bytes del archivo xlsx CONTPAq listo para importar.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pólizas"

    # ── 1. Escribir las 22 filas de encabezado ──
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    header_font = Font(bold=True, size=9)

    for row_idx, header_row in enumerate(HEADERS, start=1):
        for col_idx, valor in enumerate(header_row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.font  = header_font
            cell.fill  = header_fill

    # ── 2. Escribir una póliza por factura ──
    current_row = 23  # datos empiezan en fila 23
    num_pol     = num_poliza_inicio

    for fact in facturas:
        fecha_pol = fact["fecha"]
        folio_ref = fact["folio"]
        cta_cxc   = fact["cta_cxc"]
        total     = fact["total"]
        subtotal  = fact["subtotal"]
        iva       = fact["iva"]
        nombre    = fact["nombre_cliente"]
        uuid      = fact["uuid"]
        concepto  = f"PROVISION VENTAS {folio_ref}"

        # ── Fila P ──
        p_row = [None] * 11
        p_row[0] = "P"
        p_row[1] = fecha_pol      # datetime → CONTPAq lo leerá como fecha
        p_row[2] = int(TIPO_POL)
        p_row[3] = int(num_pol)
        p_row[4] = 1              # Clase
        p_row[5] = int(ID_DIARIO)
        p_row[6] = concepto
        p_row[7] = 12             # SistOrig (valor observado en ejemplos)
        p_row[8] = 0              # Impresa
        p_row[9] = 0              # Ajuste

        ws_row_p = current_row
        for col, val in enumerate(p_row, start=1):
            cell = ws.cell(row=ws_row_p, column=col, value=val)
            # La fecha debe ser un datetime real
            if col == 2 and isinstance(val, datetime):
                cell.number_format = "mm/dd/yyyy"
        current_row += 1

        # ── Fila M1 — CXC (Debe) ──
        m1_cxc = ["M1", int(cta_cxc), f"F-{folio_ref}", 0,
                   total, int(ID_DIARIO), 0, nombre, None]
        for col, val in enumerate(m1_cxc, start=1):
            ws.cell(row=current_row, column=col, value=val)
        current_row += 1

        # ── Fila M1 — Ingresos (Haber) ──
        m1_ing = ["M1", int(CTA_INGRESOS), f"F-{folio_ref}", 1,
                  subtotal, int(ID_DIARIO), 0, nombre, None]
        for col, val in enumerate(m1_ing, start=1):
            ws.cell(row=current_row, column=col, value=val)
        current_row += 1

        # ── Fila M1 — IVA Trasladado (Haber) ──
        if iva > 0:
            m1_iva = ["M1", int(CTA_IVA_TRAS), f"F-{folio_ref}", 1,
                      iva, int(ID_DIARIO), 0, nombre, None]
            for col, val in enumerate(m1_iva, start=1):
                ws.cell(row=current_row, column=col, value=val)
            current_row += 1

        # ── Fila AD — UUID ──
        if uuid:
            ws.cell(row=current_row, column=1, value="AD")
            ws.cell(row=current_row, column=2, value=uuid)
            current_row += 1

        num_pol += 1

    # ── Ajuste de columnas ──
    for col in ws.columns:
        max_w = 0
        for cell in col:
            if cell.value:
                max_w = max(max_w, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = min(max_w + 2, 40)

    # Guardar en memoria
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
